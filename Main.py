import sys
from datetime import timedelta
from datetime import datetime
from tkinter.font import Font

import openpyxl
import persiantools.jdatetime
from openpyxl import load_workbook
from pandas import DataFrame
import os
import os.path
from tkinter import *
from PIL import Image, ImageTk


def resource_path(relative_path):
    global base_path
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath("Assets")

    return os.path.join(base_path, relative_path)


def main_screen():
    default_font = Font(family="Mikhak Light", size=19)

    def create_excel_for_temp_file(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Date"
        sheet["A2"] = "Y/M/D"
        sheet["B1"] = "Arrived"
        sheet["B2"] = "h:m:s"
        sheet["C1"] = "BreakS"
        sheet["C2"] = "h:m:s"
        sheet["D1"] = "BreakF"
        sheet["D2"] = "h:m:s"
        sheet["E1"] = "BreakDelta"
        sheet["E2"] = 0
        sheet["F1"] = "ArriveFlag"
        sheet["F2"] = "normal"
        sheet["G1"] = "BreakTimeFlag"
        sheet["G2"] = "disabled"
        sheet["H1"] = "LeaveFlag"
        sheet["H2"] = "disabled"
        sheet["I1"] = "ThemeColor"
        sheet["I2"] = "rest_bg"
        sheet["J1"] = "ArriveBtn"
        sheet["J2"] = "arrive_btn_rest_mode"
        sheet["K1"] = "BreakBtn"
        sheet["K2"] = "leave_btn_rest_mode"
        sheet["L1"] = "LeaveBtn"
        sheet["L2"] = "break_btn_rest_mode"
        sheet["M1"] = "Leave"
        sheet["M2"] = "h:m:s"
        sheet["N1"] = "WeekDay"
        sheet["N2"] = "XDay"

        try:
            workbook.save(path)
        except:
            os.mkdir("Data Base")
            workbook.save(path)

    def create_excel_for_main_xlsx_file(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "date"
        sheet["A2"] = "Y/M/D"
        sheet["B1"] = "week day"
        sheet["B2"] = "XDay"
        sheet["C1"] = "arrive time"
        sheet["C2"] = "h:m:s"
        sheet["D1"] = "total break"
        sheet["D2"] = "h:m:s"
        sheet["E1"] = "leaving time"
        sheet["E2"] = "h:m:s"
        sheet["F1"] = "Total Working Hours"
        sheet["F2"] = "h:m:s"
        sheet["G1"] = "Working Hours till today"
        sheet["G2"] = "h:m:s"
        workbook.save(path)

    def read_from_excel(path):
        global sheet
        global all_rows

        all_excel_data = load_workbook(path)
        sheet = all_excel_data.active
        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []

        for row in rows:
            data = {}
            for all_excel_data, cell in zip(headers, row):
                data[all_excel_data] = cell.value
            all_rows.append(data)

    try:
        read_from_excel("Data Base\\Temp.xlsx")
    except:
        create_excel_for_temp_file("Data Base\\Temp.xlsx")
        read_from_excel("Data Base\\Temp.xlsx")

    def arrive():
        global weekday, arrived_date, arrived_time, weekday, bg, button_arrive, button_break, button_leave

        all_rows[0]['ThemeColor'] = "work_bg"
        all_rows[0]['ArriveBtn'] = "arrive_btn_working_mode"
        all_rows[0]['BreakBtn'] = "break_btn_working_mode"
        all_rows[0]['LeaveBtn'] = "Leave_btn_working_mode"
        all_rows[0]['ArriveFlag'] = "disabled"
        all_rows[0]['BreakTimeFlag'] = "normal"
        all_rows[0]['LeaveFlag'] = "normal"

        df = DataFrame.from_dict(all_rows)
        df.to_excel("Data Base\\Temp.xlsx")

        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\{all_rows[0]['ThemeColor']}.png")))
        bg_abel = Label(image=bg)
        bg_abel.place(x=-6, y=-2)

        button_arrive = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['ArriveBtn']}.png"))
        button_arrive_label = Label(image=button_arrive, bg="#242424")
        button_arrive_label.place(x=15, y=430)
        real_button_arrive = Button(root, image=button_arrive, borderwidth=0, bg="#242424",
                                    activebackground="#242424", state=all_rows[0]['ArriveFlag'], command=arrive)
        real_button_arrive.place(x=15, y=430)

        button_break = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['BreakBtn']}.png"))
        button_break_label = Label(image=button_break, bg="#242424")
        button_break_label.place(x=180, y=430)
        real_button_break = Button(root, image=button_break, borderwidth=0, bg="#242424",
                                   activebackground="#242424", state=all_rows[0]['BreakTimeFlag'], command=break_time)
        real_button_break.place(x=180, y=430)

        button_leave = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['LeaveBtn']}.png"))
        button_leave_label = Label(image=button_leave, bg="#242424")
        button_leave_label.place(x=340, y=430)
        real_button_leave = Button(root, image=button_leave, borderwidth=0, bg="#242424",
                                   activebackground="#242424", state=all_rows[0]['LeaveFlag'], command=leave)
        real_button_leave.place(x=340, y=430)

        try:
            read_from_excel("Data Base\\Temp.xlsx")
        except:
            create_excel_for_temp_file("Data Base\\Temp.xlsx")
            read_from_excel("Data Base\\Temp.xlsx")

        arrived_date = str(persiantools.jdatetime.JalaliDate.today().strftime("%Y/%m/%d"))
        arrived_time = str(persiantools.jdatetime.JalaliDateTime.now().time().strftime("%H:%M:%S"))
        weekday = int((persiantools.jdatetime.JalaliDate.today().strftime("%w")))

        all_rows[0]["ArriveFlag"] = "disabled"
        all_rows[0]["BreakTimeFlag"] = "normal"
        all_rows[0]["LeaveFlag"] = "normal"
        all_rows[0]["ThemeColor"] = "work_bg"

        weekday_dict = {0: "saturday", 1: "sunday", 2: "monday", 3: "tuesday", 4: "wednesday", 5: "thursday",
                        6: "friday"}

        all_rows[0]["WeekDay"] = weekday_dict[weekday]
        weekday = all_rows[0]["WeekDay"]

        all_rows[0]["Date"] = arrived_date
        all_rows[0]["Arrived"] = arrived_time
        df = DataFrame.from_dict(all_rows)
        df.to_excel("Data Base\\Temp.xlsx")

    def break_time():

        global bg, button_arrive, button_break, button_leave
        read_from_excel("Data Base\\Temp.xlsx")

        if all_rows[0]["ThemeColor"] == "rest_bg":

            all_rows[0]["ThemeColor"] = "work_bg"
            all_rows[0]['ArriveBtn'] = "arrive_btn_working_mode"
            all_rows[0]['BreakBtn'] = "break_btn_working_mode"
            all_rows[0]['LeaveBtn'] = "Leave_btn_working_mode"

            break_started_time = all_rows[0]["BreakS"]
            break_finish_time = str(persiantools.jdatetime.JalaliDateTime.now().time().strftime("%H:%M:%S"))
            all_rows[0]["BreakF"] = break_finish_time

            t1 = datetime.strptime(break_started_time, "%H:%M:%S")
            t2 = datetime.strptime(break_finish_time, "%H:%M:%S")
            delta = (t2 - t1).seconds
            delta = int(delta)

            seconds = all_rows[0]["BreakDelta"]
            seconds = int(seconds) + delta

            all_rows[0]["BreakDelta"] = seconds

            df = DataFrame.from_dict(all_rows)
            df.to_excel("Data Base\\Temp.xlsx")

            bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\{all_rows[0]['ThemeColor']}.png")))
            bg_abel = Label(image=bg)
            bg_abel.place(x=-6, y=-2)

            button_arrive = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['ArriveBtn']}.png"))
            button_arrive_label = Label(image=button_arrive, bg="#242424")
            button_arrive_label.place(x=15, y=430)
            real_button_arrive = Button(root, image=button_arrive, borderwidth=0, bg="#242424",
                                        activebackground="#242424", state=all_rows[0]['ArriveFlag'], command=arrive)
            real_button_arrive.place(x=15, y=430)

            button_break = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['BreakBtn']}.png"))
            button_break_label = Label(image=button_break, bg="#242424")
            button_break_label.place(x=180, y=430)
            real_button_break = Button(root, image=button_break, borderwidth=0, bg="#242424",
                                       activebackground="#242424", state=all_rows[0]['BreakTimeFlag'],
                                       command=break_time)
            real_button_break.place(x=180, y=430)

            button_leave = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['LeaveBtn']}.png"))
            button_leave_label = Label(image=button_leave, bg="#242424")
            button_leave_label.place(x=340, y=430)
            real_button_leave = Button(root, image=button_leave, borderwidth=0, bg="#242424",
                                       activebackground="#242424", state=all_rows[0]['LeaveFlag'], command=leave)
            real_button_leave.place(x=340, y=430)


        else:

            all_rows[0]["ThemeColor"] = "rest_bg"
            all_rows[0]['ArriveBtn'] = "arrive_btn_rest_mode"
            all_rows[0]['BreakBtn'] = "break_btn_rest_mode"
            all_rows[0]['LeaveBtn'] = "Leave_btn_rest_mode"

            break_started_time = str(persiantools.jdatetime.JalaliDateTime.now().time().strftime("%H:%M:%S"))
            all_rows[0]["BreakS"] = break_started_time

            df = DataFrame.from_dict(all_rows)
            df.to_excel("Data Base\\Temp.xlsx")

            bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\{all_rows[0]['ThemeColor']}.png")))
            bg_abel = Label(image=bg)
            bg_abel.place(x=-6, y=-2)

            button_arrive = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['ArriveBtn']}.png"))
            button_arrive_label = Label(image=button_arrive, bg="#242424")
            button_arrive_label.place(x=15, y=430)
            real_button_arrive = Button(root, image=button_arrive, borderwidth=0, bg="#242424",
                                        activebackground="#242424", state=all_rows[0]['ArriveFlag'], command=arrive)
            real_button_arrive.place(x=15, y=430)

            button_break = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['BreakBtn']}.png"))
            button_break_label = Label(image=button_break, bg="#242424")
            button_break_label.place(x=180, y=430)
            real_button_break = Button(root, image=button_break, borderwidth=0, bg="#242424",
                                       activebackground="#242424", state=all_rows[0]['BreakTimeFlag'],
                                       command=break_time)
            real_button_break.place(x=180, y=430)

            button_leave = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['LeaveBtn']}.png"))
            button_leave_label = Label(image=button_leave, bg="#242424")
            button_leave_label.place(x=340, y=430)
            real_button_leave = Button(root, image=button_leave, borderwidth=0, bg="#242424",
                                       activebackground="#242424", state=all_rows[0]['LeaveFlag'], command=leave)
            real_button_leave.place(x=340, y=430)

    def leave():

        def destroy():
            root.destroy()

        def sure_to_leave():
            global bg, button_exit, button_home

            weekday = int((persiantools.jdatetime.JalaliDate.today().strftime("%w")))
            weekday_dict = {0: "saturday", 1: "sunday", 2: "monday", 3: "tuesday", 4: "wednesday", 5: "thursday",
                            6: "friday"}
            all_rows[0]["WeekDay"] = weekday_dict[weekday]
            weekday = all_rows[0]["WeekDay"]

            now = persiantools.jdatetime.JalaliDateTime.now().time().strftime("%H:%M:%S")
            all_rows[0]["Leave"] = now
            df = DataFrame.from_dict(all_rows)
            df.to_excel("Data Base\\Temp.xlsx")

            read_from_excel("Data Base\\Temp.xlsx")

            if all_rows[0]["ThemeColor"] == "rest_bg":
                arrived_time = all_rows[0]["Arrived"]
                arrived_date = all_rows[0]["Date"]

                all_rows[0]["BreakF"] = now

                t1 = datetime.strptime(all_rows[0]["BreakS"], "%H:%M:%S")
                t2 = datetime.strptime(now, "%H:%M:%S")
                delta = (t2 - t1).seconds
                delta = int(delta)
                delta += int(all_rows[0]["BreakDelta"])
                all_rows[0]["BreakDelta"] = delta
                df = DataFrame.from_dict(all_rows)
                df.to_excel("Data Base\\Temp.xlsx")

                t1 = datetime.strptime(all_rows[0]["Arrived"], "%H:%M:%S")
                t2 = datetime.strptime(now, "%H:%M:%S")
                arrived_and_leaving_seconds_delta = (t2 - t1).seconds
                arrived_and_leaving_seconds_delta = int(arrived_and_leaving_seconds_delta)

                ideal_working_seconds = arrived_and_leaving_seconds_delta - all_rows[0]["BreakDelta"]

                minutes_get, seconds_get = divmod(ideal_working_seconds, 60)
                hours_get, minutes_get = divmod(minutes_get, 60)
                hours_get, minutes_get, seconds_get
                ideal_working_hours = f"{hours_get}:{minutes_get}:{seconds_get}"

                minutes_get, seconds_get = divmod(int(all_rows[0]["BreakDelta"]), 60)
                hours_get, minutes_get = divmod(minutes_get, 60)
                hours_get, minutes_get, seconds_get
                total_break_time = f"{hours_get}:{minutes_get}:{seconds_get}"

                try:
                    read_from_excel("Data Base\\Working Hours.xlsx")
                except:
                    create_excel_for_main_xlsx_file("Data Base\\Working Hours.xlsx")
                    read_from_excel("Data Base\\Working Hours.xlsx")

                if all_rows[-1]["date"] == "Y/M/D":
                    all_rows[-1]["week day"] = weekday
                    all_rows[-1]["arrive time"] = arrived_time
                    all_rows[-1]["date"] = arrived_date
                    all_rows[-1]["leaving time"] = str(now)
                    all_rows[-1]["total break"] = total_break_time
                    all_rows[-1]["Total Working Hours"] = ideal_working_hours

                    df = DataFrame.from_dict(all_rows)
                    df.to_excel("Data Base\\Working Hours.xlsx")
                    os.remove("Data Base\\Temp.xlsx")

                    bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\added_bg.png")))
                    bg_abel = Label(image=bg)
                    bg_abel.place(x=-6, y=-2)

                    button_exit = PhotoImage(file=resource_path(f"Images\\exit_all_clear.png"))
                    button_exit_label = Label(image=button_exit, bg="#242424")
                    button_exit_label.place(x=300, y=440)
                    real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                              activebackground="#242424", command=destroy)
                    real_button_exit.place(x=300, y=440)

                    button_home = PhotoImage(file=resource_path(f"Images\\home_all_clear.png"))
                    button_home_label = Label(image=button_home, bg="#242424")
                    button_home_label.place(x=40, y=440)
                    real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                              activebackground="#242424", command=main_screen)
                    real_button_home.place(x=40, y=440)

                else:
                    read_from_excel("Data Base\\Working Hours.xlsx")
                    all_date_list = []
                    temp = 0
                    for ______ in range(len(all_rows)):
                        all_date_list.append(all_rows[temp]["date"])
                        temp += 1

                    if arrived_date in all_date_list:

                        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\wrong_bg.png")))
                        bg_abel = Label(image=bg)
                        bg_abel.place(x=-4, y=-2)

                        button_exit = PhotoImage(file=resource_path(f"Images\\exit_sth_wrong.png"))
                        button_exit_label = Label(image=button_exit, bg="#242424")
                        button_exit_label.place(x=300, y=440)
                        real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=destroy)
                        real_button_exit.place(x=300, y=440)

                        button_home = PhotoImage(file=resource_path(f"Images\\home_sth_wrong.png"))
                        button_home_label = Label(image=button_home, bg="#242424")
                        button_home_label.place(x=40, y=440)
                        real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=main_screen)
                        real_button_home.place(x=40, y=440)

                        os.remove("Data Base\\Temp.xlsx")



                    else:
                        all_rows.append({"arrive time": arrived_time, "date": arrived_date, "week day": weekday,
                                         "leaving time": now, "total break": total_break_time,
                                         "Total Working Hours": ideal_working_hours})
                        df = DataFrame.from_dict(all_rows)
                        df.to_excel("Data Base\\Working Hours.xlsx")
                        os.remove("Data Base\\Temp.xlsx")

                        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\added_bg.png")))
                        bg_abel = Label(image=bg)
                        bg_abel.place(x=-6, y=-2)

                        button_exit = PhotoImage(file=resource_path(f"Images\\exit_all_clear.png"))
                        button_exit_label = Label(image=button_exit, bg="#242424")
                        button_exit_label.place(x=300, y=440)
                        real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=destroy)
                        real_button_exit.place(x=300, y=440)

                        button_home = PhotoImage(file=resource_path(f"Images\\home_all_clear.png"))
                        button_home_label = Label(image=button_home, bg="#242424")
                        button_home_label.place(x=40, y=440)
                        real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=main_screen)
                        real_button_home.place(x=40, y=440)

            else:
                arrived_time = all_rows[0]["Arrived"]
                arrived_date = all_rows[0]["Date"]
                try:
                    t1 = datetime.strptime(all_rows[0]["BreakS"], "%H:%M:%S")
                    t2 = datetime.strptime(all_rows[0]["BreakF"], "%H:%M:%S")
                except ValueError:
                    t1 = datetime.strptime("00:00:00", "%H:%M:%S")
                    t2 = datetime.strptime("00:00:00", "%H:%M:%S")

                delta = (t2 - t1).seconds
                delta = int(delta)
                delta += int(all_rows[0]["BreakDelta"])
                all_rows[0]["BreakDelta"] = delta
                df = DataFrame.from_dict(all_rows)
                df.to_excel("Data Base\\Temp.xlsx")

                t1 = datetime.strptime(all_rows[0]["Arrived"], "%H:%M:%S")
                t2 = datetime.strptime(now, "%H:%M:%S")
                arrived_and_leaving_seconds_delta = (t2 - t1).seconds
                arrived_and_leaving_seconds_delta = int(arrived_and_leaving_seconds_delta)

                ideal_working_seconds = arrived_and_leaving_seconds_delta - all_rows[0]["BreakDelta"]

                minutes_get, seconds_get = divmod(ideal_working_seconds, 60)
                hours_get, minutes_get = divmod(minutes_get, 60)
                hours_get, minutes_get, seconds_get
                ideal_working_hours = f"{hours_get}:{minutes_get}:{seconds_get}"

                minutes_get, seconds_get = divmod(int(all_rows[0]["BreakDelta"]), 60)
                hours_get, minutes_get = divmod(minutes_get, 60)
                hours_get, minutes_get, seconds_get
                total_break_time = f"{hours_get}:{minutes_get}:{seconds_get}"

                try:
                    read_from_excel("Data Base\\Working Hours.xlsx")
                except:
                    create_excel_for_main_xlsx_file("Data Base\\Working Hours.xlsx")
                    read_from_excel("Data Base\\Working Hours.xlsx")

                if all_rows[-1]["date"] == "Y/M/D":
                    all_rows[-1]["week day"] = weekday
                    all_rows[-1]["arrive time"] = arrived_time
                    all_rows[-1]["date"] = arrived_date
                    all_rows[-1]["leaving time"] = str(now)
                    all_rows[-1]["total break"] = total_break_time
                    all_rows[-1]["Total Working Hours"] = ideal_working_hours

                    df = DataFrame.from_dict(all_rows)
                    df.to_excel("Data Base\\Working Hours.xlsx")
                    os.remove("Data Base\\Temp.xlsx")

                    bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\added_bg.png")))
                    bg_abel = Label(image=bg)
                    bg_abel.place(x=-6, y=-2)

                    button_exit = PhotoImage(file=resource_path(f"Images\\exit_all_clear.png"))
                    button_exit_label = Label(image=button_exit, bg="#242424")
                    button_exit_label.place(x=300, y=440)
                    real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                              activebackground="#242424", command=destroy)
                    real_button_exit.place(x=300, y=440)

                    button_home = PhotoImage(file=resource_path(f"Images\\home_all_clear.png"))
                    button_home_label = Label(image=button_home, bg="#242424")
                    button_home_label.place(x=40, y=440)
                    real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                              activebackground="#242424", command=main_screen)
                    real_button_home.place(x=40, y=440)


                else:
                    read_from_excel("Data Base\\Working Hours.xlsx")
                    all_date_list = []
                    temp = 0
                    for ______ in range(len(all_rows)):
                        all_date_list.append(all_rows[temp]["date"])
                        temp += 1

                    if arrived_date in all_date_list:

                        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\wrong_bg.png")))
                        bg_abel = Label(image=bg)
                        bg_abel.place(x=-4, y=-2)

                        button_exit = PhotoImage(file=resource_path(f"Images\\exit_sth_wrong.png"))
                        button_exit_label = Label(image=button_exit, bg="#242424")
                        button_exit_label.place(x=300, y=440)
                        real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=destroy)
                        real_button_exit.place(x=300, y=440)

                        button_home = PhotoImage(file=resource_path(f"Images\\home_sth_wrong.png"))
                        button_home_label = Label(image=button_home, bg="#242424")
                        button_home_label.place(x=40, y=440)
                        real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=main_screen)
                        real_button_home.place(x=40, y=440)

                        os.remove("Data Base\\Temp.xlsx")


                    else:
                        all_rows.append({"arrive time": arrived_time, "date": arrived_date, "week day": weekday,
                                         "leaving time": now, "total break": total_break_time,
                                         "Total Working Hours": ideal_working_hours})
                        df = DataFrame.from_dict(all_rows)
                        df.to_excel("Data Base\\Working Hours.xlsx")
                        os.remove("Data Base\\Temp.xlsx")

                        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\added_bg.png")))
                        bg_abel = Label(image=bg)
                        bg_abel.place(x=-6, y=-2)

                        button_exit = PhotoImage(file=resource_path(f"Images\\exit_all_clear.png"))
                        button_exit_label = Label(image=button_exit, bg="#242424")
                        button_exit_label.place(x=300, y=440)
                        real_button_exit = Button(root, image=button_exit, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=destroy)
                        real_button_exit.place(x=300, y=440)

                        button_home = PhotoImage(file=resource_path(f"Images\\home_all_clear.png"))
                        button_home_label = Label(image=button_home, bg="#242424")
                        button_home_label.place(x=40, y=440)
                        real_button_home = Button(root, image=button_home, borderwidth=0, bg="#242424",
                                                  activebackground="#242424", command=main_screen)
                        real_button_home.place(x=40, y=440)

            read_from_excel("Data Base\\Working Hours.xlsx")

            time_list = []
            temp = 0

            for _______ in range(len(all_rows)):
                time_list.append(str(all_rows[temp]["Total Working Hours"]))
                temp += 1

            sum = timedelta()
            for ________ in time_list:
                h, m, s = ________.split(":")
                d = timedelta(hours=int(h), minutes=int(m), seconds=int(s))
                sum += d

            all_rows[-1]["Working Hours till today"] = str(sum)
            df = DataFrame.from_dict(all_rows)
            df.to_excel("Data Base\\Working Hours.xlsx")

        def change():

            def plus(min):

                if "min" in min:
                    now = persiantools.jdatetime.JalaliDateTime.now().time().strftime("%H:%M:%S")
                    now = datetime.strptime(now, "%H:%M:%S")
                    temp_arrive = datetime.strptime(all_rows[0]["Arrived"], "%H:%M:%S")

                    limit_seconds = (now - temp_arrive).seconds
                    limit_minutes = int(limit_seconds) / 60

                    how_much_to_add = min[0:-3]
                    how_much_to_add = int(how_much_to_add)
                    how_much_to_add += 5

                    how_much_to_add_to_arrive_entry.delete(0, END)
                    how_much_to_add_to_arrive_entry.insert(0, f"{how_much_to_add} min")

                    limit_check = float(str(how_much_to_add_to_arrive_entry.get()[0:-3]))

                    if limit_check > limit_minutes:
                        how_much_to_add_to_arrive_entry.delete(0, END)

                else:
                    how_much_to_add_to_arrive_entry.delete(0, END)
                    how_much_to_add_to_arrive_entry.insert(0, "0 min")

            def mines(min):

                if "min" in min:

                    how_much_to_add = min[0:-3]
                    how_much_to_add = int(how_much_to_add)
                    how_much_to_add -= 5

                    how_much_to_add_to_arrive_entry.delete(0, END)
                    how_much_to_add_to_arrive_entry.insert(0, f"{how_much_to_add} min")


                else:
                    how_much_to_add_to_arrive_entry.delete(0, END)
                    how_much_to_add_to_arrive_entry.insert(0, "0 min")

            def submit(what_to_add):
                read_from_excel("Data Base\\Temp.xlsx")

                if "min" in what_to_add:
                    what_to_add = what_to_add[0:-3]

                    all_rows[0]["Arrived"] = datetime.strptime(all_rows[0]["Arrived"], "%H:%M:%S")

                    if "-" not in what_to_add:
                        trash, all_rows[0]["Arrived"] = str(
                            all_rows[0]["Arrived"] + timedelta(minutes=int(what_to_add))).split(" ")

                    else:
                        what_to_add = what_to_add[1:]
                        what_to_add = int(what_to_add) * 60
                        trash, all_rows[0]["Arrived"] = str(
                            (all_rows[0]["Arrived"]) - timedelta(seconds=what_to_add)).split(" ")

                    df = DataFrame.from_dict(all_rows)
                    df.to_excel("Data Base\\Temp.xlsx")
                    leave()

                else:
                    how_much_to_add_to_arrive_entry.delete(0, END)
                    how_much_to_add_to_arrive_entry.insert(0, "0 min")

            global bg, button_plus, button_mines, button_submit

            bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\change_bg.png")))
            bg_abel = Label(image=bg)
            bg_abel.place(x=-6, y=-2)

            button_mines = PhotoImage(file=resource_path(f"Images\\mines_btn.png"))
            button_mines_label = Label(image=button_mines, bg="#242424")
            button_mines_label.place(x=90, y=345)
            real_button_mines = Button(root, image=button_mines, borderwidth=0, bg="#242424",
                                       activebackground="#242424",
                                       command=lambda: mines(how_much_to_add_to_arrive_entry.get()))
            real_button_mines.place(x=90, y=345)

            button_plus = PhotoImage(file=resource_path(f"Images\\plus_btn.png"))
            button_plus_label = Label(image=button_plus, bg="#242424")
            button_plus_label.place(x=340, y=345)
            real_button_plus = Button(root, image=button_plus, borderwidth=0, bg="#242424",
                                      activebackground="#242424",
                                      command=lambda: plus(how_much_to_add_to_arrive_entry.get()))
            real_button_plus.place(x=340, y=345)

            button_submit = PhotoImage(file=resource_path(f"Images\\submit_btn.png"))
            button_submit_label = Label(image=button_submit, bg="#242424")
            button_submit_label.place(x=170, y=450)
            real_button_submit = Button(root, image=button_submit, borderwidth=0, bg="#242424",
                                        activebackground="#242424",
                                        command=lambda: submit(how_much_to_add_to_arrive_entry.get()))
            real_button_submit.place(x=170, y=450)

            current_arrive_time_label = Label(root, text=all_rows[0]['Arrived'], font=default_font,
                                              bg="#4B4B4B", foreground="#ffffff")
            current_arrive_time_label.place(x=175, y=250)

            how_much_to_add_to_arrive_entry = Entry(root, borderwidth=0, font=default_font, bg="#4B4B4B",
                                                    foreground="#ffffff")
            how_much_to_add_to_arrive_entry.place(x=175, y=345, height=40, width=125)
            how_much_to_add_to_arrive_entry.insert(0, "0 min")

        global bg, button_yes, button_no, button_change

        bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\leave_bg.png")))
        bg_abel = Label(image=bg)
        bg_abel.place(x=-6, y=-2)

        button_yes = PhotoImage(file=resource_path(f"Images\\yes_btn.png"))
        button_yes_label = Label(image=button_yes, bg="#242424")
        button_yes_label.place(x=335, y=450)
        real_button_yes = Button(root, image=button_yes, borderwidth=0, bg="#242424",
                                 activebackground="#242424", command=sure_to_leave)
        real_button_yes.place(x=335, y=450)

        button_no = PhotoImage(file=resource_path(f"Images\\no_btn.png"))
        button_no_label = Label(image=button_no, bg="#242424")
        button_no_label.place(x=10, y=450)
        real_button_no = Button(root, image=button_no, borderwidth=0, bg="#242424",
                                activebackground="#242424", command=main_screen)
        real_button_no.place(x=10, y=450)

        button_change = PhotoImage(file=resource_path(f"Images\\change_btn.png"))
        button_change_label = Label(image=button_change, bg="#242424")
        button_change_label.place(x=180, y=450)
        real_button_change = Button(root, image=button_change, borderwidth=0, bg="#242424",
                                    activebackground="#242424", command=change)
        real_button_change.place(x=180, y=450)

    global bg, button_arrive, button_break, button_leave

    bg = ImageTk.PhotoImage(Image.open(resource_path(f"Images\\{all_rows[0]['ThemeColor']}.png")))
    bg_abel = Label(image=bg)
    bg_abel.place(x=-6, y=-2)

    button_arrive = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['ArriveBtn']}.png"))
    button_arrive_label = Label(image=button_arrive, bg="#242424")
    button_arrive_label.place(x=15, y=430)
    real_button_arrive = Button(root, image=button_arrive, borderwidth=0, bg="#242424",
                                activebackground="#242424", state=all_rows[0]['ArriveFlag'], command=arrive)
    real_button_arrive.place(x=15, y=430)

    button_break = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['BreakBtn']}.png"))
    button_break_label = Label(image=button_break, bg="#242424")
    button_break_label.place(x=180, y=430)
    real_button_break = Button(root, image=button_break, borderwidth=0, bg="#242424",
                               activebackground="#242424", state=all_rows[0]['BreakTimeFlag'], command=break_time)
    real_button_break.place(x=180, y=430)

    button_leave = PhotoImage(file=resource_path(f"Images\\{all_rows[0]['LeaveBtn']}.png"))
    button_leave_label = Label(image=button_leave, bg="#242424")
    button_leave_label.place(x=340, y=430)
    real_button_leave = Button(root, image=button_leave, borderwidth=0, bg="#242424",
                               activebackground="#242424", state=all_rows[0]['LeaveFlag'], command=leave)
    real_button_leave.place(x=340, y=430)


root = Tk()
root.title("PyTracker")
root.maxsize(width=475, height=550)
root.minsize(width=475, height=550)
root.iconbitmap(resource_path("Images\\clock_icon.ico"))

main_screen()

root.mainloop()
